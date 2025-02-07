from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

# Workbook = objeto python que abstrai um arquivo excel. (coleção de planilhas)
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Worksheet = objeto que abstrai uma planilha dentro do nosso arquivo excel
worksheet: Worksheet = workbook["Alunos"]

for row in worksheet.iter_rows():
    for cell in row:
        print(cell.value, end="\t")
    print()


# Alterando a idade do Mário
worksheet["B4"].value = 23

# workbook.save(WORKBOOK_PATH)
