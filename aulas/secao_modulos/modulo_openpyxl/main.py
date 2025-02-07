# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / "workbook.xlsx"

# Workbook = objeto python que abstrai um arquivo excel. (coleção de planilhas)
workbook = Workbook()
workbook.create_sheet("Alunos")
# Worksheet = objeto que abstrai uma planilha dentro do nosso arquivo excel
worksheet: Worksheet = workbook["Alunos"]
workbook.remove(workbook["Sheet"])


# Criando os cabeçalhos
"""
O método cell cria uma nova célula dentro do Worksheet (planilha) com os parâmetros especificados.
Retorna a célula recém criada logo em seguida.
"""
worksheet.cell(1, 1, "Nome")
worksheet.cell(1, 2, "Idade")
worksheet.cell(1, 3, "Nota")

students = [
    ["Gustavo", 20, 10],
    ["Isabela", 19, 7.5],
    ["Mario", 20, 8],
]

# for line_index in range(2, len(students) + 2):
#     for colum_index in range(1, 4):
#         worksheet.cell(
#             line_index, colum_index, students[line_index - 2][colum_index - 1]
#         )

for student in students:
    worksheet.append(student)


workbook.save(WORKBOOK_PATH)
